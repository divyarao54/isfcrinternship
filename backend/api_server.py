from flask import Flask, request, jsonify
#from tasks import scrape_papers_task, cleanup_old_jobs_task, health_check_task
#from scheduler_python import scheduler
import os
from dotenv import load_dotenv
import subprocess
from flask_cors import CORS
import urllib.parse
import re
from datetime import datetime
import json
from bson.objectid import ObjectId
from werkzeug.utils import secure_filename
import bcrypt
import secrets
import string
import threading
import time
import uuid
from db_config import db_manager
import io
import csv
from openpyxl import load_workbook

load_dotenv()

app = Flask(__name__)
CORS(app)

@app.route('/health', methods=['GET'])
def health_check():
	return jsonify({'status': 'ok'}), 200
# -------- Awards config ---------
def get_frontend_assets_dir():
    backend_dir = os.path.dirname(__file__)
    project_root = os.path.abspath(os.path.join(backend_dir, '..'))
    assets_dir = os.path.join(project_root, 'frontend', 'src', 'assets')
    os.makedirs(assets_dir, exist_ok=True)
    return assets_dir

def slugify_filename(name: str) -> str:
    # Keep alphanumerics, replace spaces with underscores, drop others
    import re
    s = name.strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_\-]", "", s)
    return s or 'award'

@app.route('/api/awards', methods=['POST'])
def add_award():
    """Add an ISFCR award with image URL; stores record in MongoDB awards collection."""
    try:
        print('[AWARDS] Handling add_award')
        
        # Get JSON data instead of form data
        data = request.get_json()
        if not data:
            return jsonify({'error': 'JSON data is required'}), 400
            
        # Validate fields
        award_name = data.get('awardName', '').strip()
        image_url = data.get('imageUrl', '').strip()
        print(f'[AWARDS] awardName={award_name!r}')
        print(f'[AWARDS] imageUrl={image_url!r}')
        
        if not award_name:
            return jsonify({'error': 'Award name is required'}), 400
        if not image_url:
            return jsonify({'error': 'Image URL is required'}), 400
            
        # Basic URL validation
        try:
            from urllib.parse import urlparse
            parsed_url = urlparse(image_url)
            if not parsed_url.scheme or not parsed_url.netloc:
                return jsonify({'error': 'Invalid image URL format'}), 400
        except Exception:
            return jsonify({'error': 'Invalid image URL format'}), 400

        # Store metadata in DB
        db = db_manager.connect()
        awards_collection = db['awards']
        doc = {
            'awardName': award_name,
            'imageUrl': image_url,
            'createdAt': datetime.utcnow().isoformat()
        }
        result = awards_collection.insert_one(doc)

        print(f'[AWARDS] inserted id={str(result.inserted_id)}')
        # Build a JSON-safe award payload (avoid ObjectId in doc after insert_one)
        safe_award = {k: v for k, v in doc.items() if k != '_id'}
        safe_award['_id'] = str(result.inserted_id)
        return jsonify({
            'success': True,
            'message': 'Award added successfully',
            'award': safe_award
        }), 201
    except Exception as e:
        import traceback
        print('[AWARDS] ERROR:', str(e))
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/awards', methods=['GET'])
def list_awards():
    """List uploaded ISFCR awards."""
    try:
        db = db_manager.connect()
        awards_collection = db['awards']
        awards = list(awards_collection.find({}).sort([('createdAt', -1)]))
        for a in awards:
            a['_id'] = str(a['_id'])
        return jsonify({'success': True, 'awards': awards}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Task management for background processes
task_store = {}
task_lock = threading.Lock()

def get_task_status(task_id):
    """Get the status of a background task"""
    with task_lock:
        return task_store.get(task_id, {'status': 'not_found'})

def update_task_status(task_id, status, result=None, error=None):
    """Update the status of a background task"""
    with task_lock:
        if task_id in task_store:
            task_store[task_id].update({
                'status': status,
                'result': result,
                'error': error,
                'updated_at': datetime.now().isoformat()
            })

def create_task(task_type, params):
    """Create a new background task"""
    task_id = str(uuid.uuid4())
    with task_lock:
        task_store[task_id] = {
            'id': task_id,
            'type': task_type,
            'params': params,
            'status': 'pending',
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }
    return task_id

def run_scraping_task(task_id, profile_url):
    """Run scraping in background"""
    try:
        update_task_status(task_id, 'running')
        
        # Run the scraping process
        import subprocess
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as stdout_file, \
             tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as stderr_file:
            
            stdout_path = stdout_file.name
            stderr_path = stderr_file.name
        
        scraper_process = subprocess.Popen(
            ['node', 'scraper.js', profile_url],
            cwd=os.path.join(os.path.dirname(__file__)),
            stdout=open(stdout_path, 'w', encoding='utf-8'),
            stderr=open(stderr_path, 'w', encoding='utf-8'),
            env=dict(os.environ, PYTHONIOENCODING='utf-8', NODE_OPTIONS='--max-old-space-size=4096')
        )
        
        return_code = scraper_process.wait(timeout=3600)
        
        with open(stdout_path, 'r', encoding='utf-8', errors='replace') as f:
            stdout = f.read()
        with open(stderr_path, 'r', encoding='utf-8', errors='replace') as f:
            stderr = f.read()
        
        # Clean up temp files
        try:
            os.unlink(stdout_path)
            os.unlink(stderr_path)
        except:
            pass
        
        if return_code == 0:
            update_task_status(task_id, 'completed', {'message': 'Scraping completed successfully'})
        else:
            update_task_status(task_id, 'failed', error=stderr)
            
    except Exception as e:
        update_task_status(task_id, 'failed', error=str(e))

def run_update_all_task(task_id):
    """Run update all in background"""
    try:
        update_task_status(task_id, 'running')
        
        # Run the update all process
        from db_config import get_collection
        import subprocess, tempfile, os
        
        teachers_collection = get_collection('teachers')
        teachers = list(teachers_collection.find({}, {'profileUrl': 1, 'name': 1}))
        
        if not teachers:
            update_task_status(task_id, 'failed', error='No teachers found')
            return
        
        results = []
        for idx, teacher in enumerate(teachers, 1):
            profile_url = teacher.get('profileUrl')
            name = teacher.get('name')
            if not profile_url:
                continue
                
            update_task_status(task_id, 'running', {'message': f'Updating {name} ({idx}/{len(teachers)})'})
            
            with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as stdout_file, \
                 tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as stderr_file:
                stdout_path = stdout_file.name
                stderr_path = stderr_file.name
            
            scraper_process = subprocess.Popen(
                ['node', 'scraper.js', profile_url],
                cwd=os.path.join(os.path.dirname(__file__)),
                stdout=open(stdout_path, 'w', encoding='utf-8'),
                stderr=open(stderr_path, 'w', encoding='utf-8'),
                env=dict(os.environ, PYTHONIOENCODING='utf-8', NODE_OPTIONS='--max-old-space-size=4096')
            )
            
            try:
                return_code = scraper_process.wait(timeout=3600)
            except subprocess.TimeoutExpired:
                scraper_process.kill()
                return_code = -1
            
            with open(stdout_path, 'r', encoding='utf-8', errors='replace') as f:
                stdout = f.read()
            with open(stderr_path, 'r', encoding='utf-8', errors='replace') as f:
                stderr = f.read()
            
            results.append({
                'teacher': name,
                'profileUrl': profile_url,
                'return_code': return_code,
                'stdout': stdout[:500],
                'stderr': stderr[:500]
            })
            
            # Clean up temp files
            try:
                os.unlink(stdout_path)
                os.unlink(stderr_path)
            except:
                pass
        
        # Complete task (Elasticsearch sync removed)
        update_task_status(task_id, 'completed', {
            'message': f'Update completed for {len(teachers)} teachers',
            'results': results
        })
        
    except Exception as e:
        update_task_status(task_id, 'failed', error=str(e))

# Background task endpoints
@app.route('/tasks/start-scraping', methods=['POST'])
def start_scraping_task():
    """Start a background scraping task"""
    try:
        data = request.get_json()
        profile_url = data.get('profileUrl')
        if not profile_url:
            return jsonify({'error': 'No profileUrl provided'}), 400
        
        task_id = create_task('scraping', {'profileUrl': profile_url})
        
        # Start the task in a background thread
        thread = threading.Thread(target=run_scraping_task, args=(task_id, profile_url))
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'task_id': task_id,
            'status': 'started',
            'message': 'Scraping task started in background'
        }), 200
        
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/tasks/start-update-all', methods=['POST'])
def start_update_all_task():
    """Start a background update all task"""
    try:
        task_id = create_task('update_all', {})
        
        # Start the task in a background thread
        thread = threading.Thread(target=run_update_all_task, args=(task_id,))
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'task_id': task_id,
            'status': 'started',
            'message': 'Update all task started in background'
        }), 200
        
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/tasks/<task_id>/status', methods=['GET'])
def get_task_status_endpoint(task_id):
    """Get the status of a background task"""
    try:
        status = get_task_status(task_id)
        return jsonify(status), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500



@app.route('/tasks', methods=['GET'])
def list_tasks():
    """List all background tasks"""
    try:
        with task_lock:
            tasks = list(task_store.values())
        return jsonify({'tasks': tasks}), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/papers/count', methods=['GET'])
def get_papers_count():
    """Get count of papers in database from all teacher collections"""
    try:
        from db_config import get_collection
        db = get_collection('teachers').database
        all_collections = db.list_collection_names()
        teacher_paper_collections = [col for col in all_collections if col.startswith('papers_')]
        
        total_count = 0
        for col_name in teacher_paper_collections:
            col = db.get_collection(col_name)
            count = col.count_documents({})
            total_count += count
        
        return jsonify({'count': total_count}), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/papers', methods=['GET'])
def get_papers():
    """Get papers from database from all teacher collections"""
    try:
        from db_config import get_collection
        db = get_collection('teachers').database
        all_collections = db.list_collection_names()
        teacher_paper_collections = [col for col in all_collections if col.startswith('papers_')]
        
        # Get query parameters
        limit = int(request.args.get('limit', 50))
        skip = int(request.args.get('skip', 0))
        
        # Gather all papers from all teacher collections
        all_papers = []
        for col_name in teacher_paper_collections:
            col = db.get_collection(col_name)
            papers = list(col.find({}, {'_id': 0}))
            all_papers.extend(papers)
        
        # Apply pagination
        total_papers = len(all_papers)
        paginated_papers = all_papers[skip:skip + limit]
        
        return jsonify({
            'papers': paginated_papers,
            'total': total_papers,
            'limit': limit,
            'skip': skip
        }), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/teachers', methods=['GET'])
def get_teachers():
    """Get all teachers from database"""
    try:
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        
        # Get all teachers, including _id, and convert _id to string
        teachers = list(teachers_collection.find({}))
        for teacher in teachers:
            if '_id' in teacher:
                teacher['_id'] = str(teacher['_id'])
        
        return jsonify({
            'teachers': teachers,
            'total': len(teachers)
        }), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/teachers/debug', methods=['GET'])
def debug_teachers():
    """Debug endpoint to see all teachers in database"""
    try:
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        
        # Get all teachers with their names
        teachers = list(teachers_collection.find({}, {'name': 1, 'profileUrl': 1, '_id': 0}))
        
        return jsonify({
            'teachers': teachers,
            'total': len(teachers)
        }), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/teachers/<teacher_id>', methods=['GET'])
def get_teacher(teacher_id):
    """Get individual teacher by name"""
    try:
        print(f"=== GET /teachers/{teacher_id} ===")
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        print(f"Looking for teacher with name: '{decoded_teacher_id}'")
        teacher = teachers_collection.find_one({'name': decoded_teacher_id}, {'_id': 0})
        if not teacher:
            all_teachers = list(teachers_collection.find({}, {'name': 1, '_id': 0}))
            print(f"Available teachers: {[t['name'] for t in all_teachers]}")
            print(f"404: Teacher '{decoded_teacher_id}' not found")
            return jsonify({'error': 'Teacher not found'}), 404
        print(f"Found teacher: {teacher['name']}")
        print(f"Returning teacher data: {teacher}")
        return jsonify({'teacher': teacher}), 200
    except Exception as error:
        print(f"Error in get_teacher: {error}")
        return jsonify({'error': str(error)}), 500

@app.route('/teachers/<teacher_id>/stats', methods=['GET'])
def get_teacher_stats(teacher_id):
    """Get statistics for an individual teacher (journal vs. conference papers)."""
    try:
        print(f"=== GET /teachers/{teacher_id}/stats ===")
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        print(f"Looking for teacher with name: '{decoded_teacher_id}'")
        teacher = teachers_collection.find_one({'name': decoded_teacher_id})
        if not teacher:
            all_teachers = list(teachers_collection.find({}, {'name': 1, '_id': 0}))
            print(f"Available teachers: {[t['name'] for t in all_teachers]}")
            print(f"404: Teacher '{decoded_teacher_id}' not found")
            return jsonify({'error': 'Teacher not found'}), 404
        teacher_name = teacher['name']
        collection_name = 'papers_' + re.sub(r'[^a-z0-9]', '_', teacher_name.lower())
        print(f"Using collection: {collection_name}")
        papers_collection = get_collection(collection_name)
        # Count journals: papers with non-empty journal field OR non-empty source field
        journal_count = papers_collection.count_documents({
            '$or': [
                {'journal': {'$exists': True, '$ne': ''}},
                {'source': {'$exists': True, '$ne': ''}}
            ]
        })
        conference_count = papers_collection.count_documents({'conference': {'$exists': True, '$ne': ''}})
        book_count = papers_collection.count_documents({'book': {'$exists': True, '$ne': ''}})
        # Count patents with more robust query to handle different boolean representations
        patent_count = papers_collection.count_documents({
            '$or': [
                {'patent': True},
                {'patent': 'true'},
                {'patent': 'True'}
            ]
        })
        print(f"Journal count: {journal_count}, Conference count: {conference_count}, Book count: {book_count}, Patent count: {patent_count}")
        return jsonify({
            'journal_count': journal_count,
            'conference_count': conference_count,
            'book_count': book_count,
            'patent_count': patent_count
        }), 200
    except Exception as e:
        print(f"Error in get_teacher_stats: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/teachers/<teacher_id>/citations/scholar', methods=['GET'])
def get_scholar_citations(teacher_id):
    """Get citation statistics from Google Scholar profile using the new scraper."""
    try:
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        teacher = teachers_collection.find_one({'name': decoded_teacher_id})
        if not teacher:
            return jsonify({'error': 'Teacher not found'}), 404
        
        profile_url = teacher.get('profileUrl')
        if not profile_url:
            return jsonify({'error': 'Teacher profile URL not found'}), 404
            
        script_path = os.path.join(os.path.dirname(__file__), 'scholarCitationsPerYear.js')
        if not os.path.exists(script_path):
             return jsonify({'error': 'scholarCitationsPerYear.js not found'}), 500

        result = subprocess.run(
            ['node', script_path, profile_url],
            capture_output=True, text=True, timeout=120 # 2 minute timeout
        )
        
        if result.returncode != 0:
            print(f"Error from scholarCitationsPerYear.js: {result.stderr}")
            return jsonify({'error': 'Failed to fetch citation data from Scholar', 'details': result.stderr}), 500
        
        try:
            # The output from the script is a JSON string
            citations = json.loads(result.stdout)
            return jsonify({'citations': citations}), 200
        except json.JSONDecodeError:
            print(f"Failed to parse JSON from scholar scraper: {result.stdout}")
            return jsonify({'error': 'Invalid data format from scholar scraper'}), 500
            
    except subprocess.TimeoutExpired:
        return jsonify({'error': 'Scraping Google Scholar timed out'}), 500
    except Exception as error:
        print(f"Server error in /citations/scholar: {error}")
        return jsonify({'error': str(error)}), 500

@app.route('/teachers/<teacher_id>/citations', methods=['GET'])
def get_teacher_citations(teacher_id):
    """Get citation statistics from Google Scholar profile"""
    try:
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        
        # Decode the URL-encoded teacher_id
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        
        # Find teacher by name first
        teacher = teachers_collection.find_one({'name': decoded_teacher_id})
        if not teacher:
            return jsonify({'error': 'Teacher not found'}), 404
        
        # Use the teacher's profileUrl for scraping
        profile_url = teacher['profileUrl']
        
        # Import the scraper function
        import subprocess
        import json
        
        # Call a Node.js script to scrape citation data
        result = subprocess.run(
            ['node', 'citation_scraper.js', profile_url],
            cwd=os.path.join(os.path.dirname(__file__)),
            capture_output=True, text=True, timeout=60
        )
        
        if result.returncode != 0:
            return jsonify({'error': 'Failed to fetch citation data'}), 500
        
        try:
            citations = json.loads(result.stdout)
            return jsonify({'citations': citations}), 200
        except json.JSONDecodeError:
            return jsonify({'error': 'Invalid citation data format'}), 500
            
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/admin/scrape', methods=['POST'])
def admin_scrape():
    """Start background scraping - returns immediately with task ID"""
    return start_scraping_task()

@app.route('/admin/update-all', methods=['POST'])
def admin_update_all():
    """Start background update all - returns immediately with task ID"""
    return start_update_all_task()

@app.route('/teachers/<teacher_id>/publications', methods=['GET'])
def get_teacher_publications(teacher_id):
    """Get all publications for a teacher, sorted by number of citations (descending)"""
    try:
        print(f"=== GET /teachers/{teacher_id}/publications ===")
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        import urllib.parse
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        print(f"Looking for teacher with name: '{decoded_teacher_id}'")
        teacher = teachers_collection.find_one({'name': decoded_teacher_id})
        if not teacher:
            all_teachers = list(teachers_collection.find({}, {'name': 1, '_id': 0}))
            print(f"Available teachers: {[t['name'] for t in all_teachers]}")
            print(f"404: Teacher '{decoded_teacher_id}' not found")
            return jsonify({'error': 'Teacher not found'}), 404
        teacher_name = teacher['name']
        collection_name = 'papers_' + re.sub(r'[^a-z0-9]', '_', teacher_name.lower())
        print(f"Using collection: {collection_name}")
        teacher_papers_collection = get_collection(collection_name)
        # Do not exclude _id so it is included in the response
        papers = list(teacher_papers_collection.find({}).sort('citationCount', -1))
        # Convert _id to string for each paper
        for paper in papers:
            if '_id' in paper:
                paper['_id'] = str(paper['_id'])
        print(f"Found {len(papers)} papers for teacher {teacher_name}")
        return jsonify({'publications': papers, 'total': len(papers)}), 200
    except Exception as error:
        print(f"Error in get_teacher_publications: {error}")
        return jsonify({'error': str(error)}), 500

@app.route('/teachers/<teacher_id>/publications/<publication_id>', methods=['GET'])
def get_publication_details(teacher_id, publication_id):
    """Get specific publication details by publication title"""
    try:
        print(f"=== GET /teachers/{teacher_id}/publications/{publication_id} ===")
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        
        # Decode the teacher_id
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        print(f"Looking for teacher: '{decoded_teacher_id}'")
        teacher = teachers_collection.find_one({'name': decoded_teacher_id})
        if not teacher:
            print(f"Teacher '{decoded_teacher_id}' not found")
            return jsonify({'error': 'Teacher not found'}), 404
        
        # Decode the publication_id (which is the URL-encoded publication title)
        decoded_publication_title = urllib.parse.unquote(publication_id)
        print(f"Looking for publication: '{decoded_publication_title}'")
        
        # Get the teacher's specific papers collection
        teacher_name = teacher['name']
        # Sanitize collection name (same logic as in scraper.js)
        collection_name = 'papers_' + re.sub(r'[^a-z0-9]', '_', teacher_name.lower())
        print(f"Using collection: '{collection_name}'")
        
        # Get the specific collection for this teacher
        teacher_papers_collection = get_collection(collection_name)
        
        # Find the specific publication by title
        publication = teacher_papers_collection.find_one(
            {'title': decoded_publication_title}, 
            {'_id': 0}
        )
        
        if not publication:
            # Debug: list all publications in this collection
            all_publications = list(teacher_papers_collection.find({}, {'title': 1, '_id': 0}))
            print(f"Available publications in {collection_name}: {[p['title'] for p in all_publications]}")
            print(f"404: Publication '{decoded_publication_title}' not found")
            return jsonify({'error': 'Publication not found'}), 404
        
        print(f"Found publication: {publication['title']}")
        return jsonify({'publication': publication}), 200
    except Exception as error:
        print(f"Error in get_publication_details: {error}")
        return jsonify({'error': str(error)}), 500

@app.route('/teachers/<teacher_id>/publications/<pub_id>', methods=['DELETE'])
def delete_teacher_publication(teacher_id, pub_id):
    """Delete a publication by ObjectId from a teacher's papers collection"""
    try:
        from db_config import get_collection
        import re, urllib.parse
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        collection_name = 'papers_' + re.sub(r'[^a-z0-9]', '_', decoded_teacher_id.lower())
        papers_collection = get_collection(collection_name)
        result = papers_collection.delete_one({'_id': ObjectId(pub_id)})
        if result.deleted_count == 1:
            return jsonify({'message': 'Publication deleted'}), 200
        else:
            return jsonify({'error': 'Publication not found'}), 404
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/elasticsearch/status', methods=['GET'])
def check_elasticsearch_status():
	return jsonify({'status': 'disabled', 'message': 'Elasticsearch is no longer used. Search is powered by MongoDB.'}), 200

@app.route('/search', methods=['GET'])
def search_papers():
    """Search papers across all MongoDB teacher collections using case-insensitive regex."""
    try:
        from db_config import get_collection
        import re as _re

        query = request.args.get('q', '').strip()
        if not query:
            return jsonify({'error': 'No search query provided'}), 400
        try:
            limit = int(request.args.get('limit', 50))
        except Exception:
            limit = 50

        escaped = _re.escape(query)
        regex = {'$regex': escaped, '$options': 'i'}

        db = get_collection('teachers').database
        all_collections = db.list_collection_names()
        teacher_paper_collections = [col for col in all_collections if col.startswith('papers_')]

        results = []
        for col_name in teacher_paper_collections:
            col = db.get_collection(col_name)
            cursor = col.find({
                '$or': [
                    {'title': regex},
                    {'authors': regex},
                    {'description': regex},
                    {'summary': regex},
                ]
            })

            for doc in cursor:
                results.append({
                    '_id': str(doc.get('_id', '')),
                    'title': doc.get('title', ''),
                    'authors': doc.get('authors', ''),
                    'year': doc.get('year', ''),
                    'url': doc.get('url', ''),
                    'teacherName': doc.get('teacherName', ''),
                    'index': col_name,
                    'citationCount': doc.get('citationCount', 0),
                    'description': doc.get('description', ''),
                    'summary': doc.get('summary', ''),
                    'source': doc.get('source', ''),
                    'journal': doc.get('journal', ''),
                    'conference': doc.get('conference', ''),
                    'book': doc.get('book', ''),
                    'pdfLink': doc.get('pdfLink', ''),
                })

        seen_keys = set()
        unique_results = []
        for hit in results:
            raw_title = hit.get('title', '')
            if not isinstance(raw_title, str):
                raw_title = str(raw_title or '')
            title = raw_title.strip().lower()

            year = hit.get('year') or ''

            raw_authors = hit.get('authors', '')
            if not isinstance(raw_authors, str):
                raw_authors = str(raw_authors or '')
            authors = raw_authors.strip().lower()

            key = f"{title}|{year}|{authors}"
            if key in seen_keys:
                continue
            seen_keys.add(key)
            unique_results.append(hit)

        final_results = unique_results[:max(1, limit)]

        return jsonify({
            'query': query,
            'total_results': len(final_results),
            'results': final_results
        }), 200
            
    except Exception as error:
        return jsonify({'error': 'Error performing search', 'message': str(error)}), 500

@app.route('/api/community/stats', methods=['GET'])
def get_community_stats():
    """Get aggregated community-wide statistics from all teacher paper collections."""
    try:
        from db_config import get_collection
        db = get_collection('teachers').database
        all_collections = db.list_collection_names()
        teacher_paper_collections = [col for col in all_collections if col.startswith('papers_')]
        
        # Gather all papers from all teacher collections
        all_papers = []
        for col_name in teacher_paper_collections:
            col = db.get_collection(col_name)
            papers = list(col.find({}))
            all_papers.extend(papers)
        
        # De-duplicate papers by URL before counting
        papers_by_url = {}
        for paper in all_papers:
            url = paper.get('url')
            if url and url not in papers_by_url:
                papers_by_url[url] = paper
        unique_papers = list(papers_by_url.values())
        
        journal_count = 0
        conference_count = 0
        book_count = 0
        patent_count = 0
        
        for paper in unique_papers:
            if (paper.get('journal') or paper.get('source')):
                journal_count += 1
            if paper.get('conference'):
                conference_count += 1
            if paper.get('book'):
                book_count += 1
            if paper.get('patent') is True or paper.get('patent') == True or (isinstance(paper.get('patent'), str) and paper.get('patent').lower() == 'true'):
                patent_count += 1
        
        return jsonify({
            'journal_count': journal_count,
            'conference_count': conference_count,
            'book_count': book_count,
            'patent_count': patent_count
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/community/yearly_stats', methods=['GET'])
def get_community_yearly_stats():
    """Aggregate year-wise publication, journal, conference, book, and patent counts from all teachers' papers collections."""
    try:
        from db_config import get_collection
        import re
        from datetime import datetime
        import pymongo

        debug_info = []

        # Get all teacher collections
        db = get_collection('teachers').database
        all_collections = db.list_collection_names()
        teacher_paper_collections = [col for col in all_collections if col.startswith('papers_') and col != 'papers']

        debug_info.append(f"Found {len(teacher_paper_collections)} teacher paper collections")

        # Gather all papers from all teacher collections
        all_papers = []
        for col_name in teacher_paper_collections:
            col = db.get_collection(col_name)
            papers = list(col.find({}))
            debug_info.append(f"Collection {col_name}: {len(papers)} papers")
            all_papers.extend(papers)

        debug_info.append(f"Total papers before deduplication: {len(all_papers)}")

        # De-duplicate by URL
        papers_by_url = {}
        for paper in all_papers:
            url = paper.get('url')
            if url and url not in papers_by_url:
                papers_by_url[url] = paper
        unique_papers = list(papers_by_url.values())

        debug_info.append(f"Unique papers after deduplication: {len(unique_papers)}")

        yearly_data = {}
        books_count = 0
        conferences_count = 0
        journals_count = 0
        patents_count = 0
        
        book_examples = []
        journal_examples = []
        conference_examples = []
        
        def extract_year_from_paper(paper_doc):
            year_value = paper_doc.get('year')
            # Handle numeric year
            try:
                if year_value is not None and str(int(year_value)).isdigit():
                    return int(year_value)
            except Exception:
                pass

            # Try common date fields
            for date_key in ['publicationDate', 'grantedOn', 'filedOn']:
                date_val = paper_doc.get(date_key)
                if date_val and isinstance(date_val, str) and len(date_val) >= 4:
                    first_four = date_val[:4]
                    if first_four.isdigit():
                        return int(first_four)
            return None

        for paper in unique_papers:
            year = extract_year_from_paper(paper)
            if year is None:
                continue
            if year not in yearly_data:
                yearly_data[year] = {
                    'conferences': 0, 'journals': 0, 'books': 0, 'patents': 0, 'papers': 0
                }
            yearly_data[year]['papers'] += 1

            # Categorize the paper
            source = paper.get('source')
            journal = paper.get('journal')
            conference = paper.get('conference')
            book = paper.get('book')
            patent = paper.get('patent')

            # Check for books first (non-empty book field)
            if book:
                yearly_data[year]['books'] += 1
                books_count += 1
                if len(book_examples) < 3:
                    book_examples.append({
                        'title': paper.get('title', 'No title')[:100],
                        'book': str(book),
                        'source': str(source) if source else '',
                        'journal': str(journal) if journal else '',
                        'conference': str(conference) if conference else ''
                    })
            # Check for conferences (non-empty conference field)
            elif conference:
                yearly_data[year]['conferences'] += 1
                conferences_count += 1
                if len(conference_examples) < 3:
                    conference_examples.append({
                        'title': paper.get('title', 'No title')[:100],
                        'conference': str(conference),
                        'source': str(source) if source else '',
                        'journal': str(journal) if journal else '',
                        'book': str(book) if book else ''
                    })
            # Check for patents (patent field True or "true")
            elif patent is True or patent == True or (isinstance(patent, str) and patent.lower() == 'true'):
                yearly_data[year]['patents'] += 1
                patents_count += 1
            # Check for journals (non-empty source field OR non-empty journal field)
            elif source or journal:
                yearly_data[year]['journals'] += 1
                journals_count += 1
                if len(journal_examples) < 3:
                    journal_examples.append({
                        'title': paper.get('title', 'No title')[:100],
                        'source': str(source) if source else '',
                        'journal': str(journal) if journal else '',
                        'conference': str(conference) if conference else '',
                        'book': str(book) if book else ''
                    })

        debug_info.append(f"Final counts - Books: {books_count}, Conferences: {conferences_count}, Journals: {journals_count}, Patents: {patents_count}")

        # Convert to list of objects for the frontend, sorted by year
        yearly_stats = [{'year': y, **data} for y, data in sorted(yearly_data.items())]

        # Compute summary
        total_publications = sum(y['papers'] for y in yearly_stats)
        total_conferences = sum(y['conferences'] for y in yearly_stats)
        total_journals = sum(y['journals'] for y in yearly_stats)
        total_books = sum(y['books'] for y in yearly_stats)
        total_patents = sum(y['patents'] for y in yearly_stats)
        
        summary = {
            'total_publications': total_publications,
            'total_conferences': total_conferences,
            'total_journals': total_journals,
            'total_books': total_books,
            'total_patents': total_patents
        }

        debug_info.append(f"Summary object: {summary}")

        return jsonify({
            'summary': summary,
            'yearly_stats': yearly_stats,
            'debug_info': debug_info,
            'examples': {
                'books': book_examples,
                'journals': journal_examples,
                'conferences': conference_examples
            }
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/teachers/<teacher_id>/citations/stored', methods=['GET'])
def get_stored_citations_by_year(teacher_id):
    """Aggregate total citationCount by publication year for a teacher from their papers collection."""
    try:
        print(f"=== GET /teachers/{teacher_id}/citations/stored ===")
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        import urllib.parse, re
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        print(f"Looking for teacher with name: '{decoded_teacher_id}'")
        teacher = teachers_collection.find_one({'name': decoded_teacher_id})
        if not teacher:
            all_teachers = list(teachers_collection.find({}, {'name': 1, '_id': 0}))
            print(f"Available teachers: {[t['name'] for t in all_teachers]}")
            print(f"404: Teacher '{decoded_teacher_id}' not found")
            return jsonify({'error': 'Teacher not found'}), 404
        teacher_name = teacher['name']
        collection_name = 'papers_' + re.sub(r'[^a-z0-9]', '_', teacher_name.lower())
        print(f"Using collection: {collection_name}")
        papers_collection = get_collection(collection_name)
        citations_per_year = {}
        for paper in papers_collection.find({}):
            year = paper.get('year')
            count = paper.get('citationCount', 0)
            if year and str(year).isdigit():
                citations_per_year[year] = citations_per_year.get(year, 0) + int(count)
        sorted_citations = dict(sorted(citations_per_year.items(), key=lambda x: x[0]))
        print(f"Citations per year: {sorted_citations}")
        return jsonify({
            'teacherName': teacher_name,
            'citationsPerYear': sorted_citations
        }), 200
    except Exception as error:
        print(f"Error in get_stored_citations_by_year: {error}")
        return jsonify({'error': str(error)}), 500

@app.route('/teachers/<teacher_id>/citations/scraped', methods=['GET'])
def get_scraped_citations_by_year(teacher_id):
    """Get citations by year for a teacher from the citations collection (scraped from Google Scholar)."""
    try:
        print(f"=== GET /teachers/{teacher_id}/citations/scraped ===")
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        citations_collection = get_collection('citations')
        import urllib.parse
        
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        print(f"Looking for teacher with name: '{decoded_teacher_id}'")
        teacher = teachers_collection.find_one({'name': decoded_teacher_id})
        
        if not teacher:
            all_teachers = list(teachers_collection.find({}, {'name': 1, '_id': 0}))
            print(f"Available teachers: {[t['name'] for t in all_teachers]}")
            print(f"404: Teacher '{decoded_teacher_id}' not found")
            return jsonify({'error': 'Teacher not found'}), 404
        
        # Get citations from citations collection using teacherName
        citation_doc = citations_collection.find_one({'teacherName': teacher['name']})
        
        if not citation_doc:
            print(f"No citation data found for teacher {teacher['name']}")
            return jsonify({
                'teacherName': teacher['name'],
                'citationsPerYear': {},
                'lastUpdated': None,
                'totalCitations': 0,
                'hIndex': 0,
                'i10Index': 0
            }), 200
        
        # Convert Map to dict if needed
        citations_per_year = citation_doc.get('citationsPerYear', {})
        if hasattr(citations_per_year, 'to_dict'):
            citations_per_year = citations_per_year.to_dict()
        
        # Calculate total citations
        total_citations = sum(v for v in citations_per_year.values() if isinstance(v, (int, float)))
        h_index = citation_doc.get('hIndex', 0)
        i10_index = citation_doc.get('i10Index', 0)
        
        print(f"Found citation data for {teacher['name']}: {citations_per_year}")
        return jsonify({
            'teacherName': citation_doc['teacherName'],
            'citationsPerYear': citations_per_year,
            'lastUpdated': citation_doc.get('lastUpdated'),
            'totalCitations': total_citations,
            'hIndex': h_index,
            'i10Index': i10_index
        }), 200
        
    except Exception as error:
        print(f"Error in get_scraped_citations_by_year: {error}")
        return jsonify({'error': str(error)}), 500

@app.route('/teachers/<teacher_id>', methods=['DELETE'])
def delete_teacher(teacher_id):
    """Delete a teacher and all their related data."""
    try:
        from db_config import get_collection
        teachers_collection = get_collection('teachers')
        # Decode the teacher_id (assumed to be the teacher's name)
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        teacher = teachers_collection.find_one({'name': decoded_teacher_id})
        if not teacher:
            return jsonify({'error': 'Teacher not found'}), 404

        # Remove teacher
        teachers_collection.delete_one({'name': decoded_teacher_id})

        # Remove related papers collection
        import re
        collection_name = 'papers_' + re.sub(r'[^a-z0-9]', '_', decoded_teacher_id.lower())
        db = teachers_collection.database
        if collection_name in db.list_collection_names():
            db.drop_collection(collection_name)

        # Remove related citations
        citations_collection = get_collection('citations')
        citations_collection.delete_many({'teacherName': decoded_teacher_id})

        return jsonify({'message': f"Teacher '{decoded_teacher_id}' and related data deleted."}), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/api/community/update_total_citations', methods=['POST'])
def update_total_citations():
    """Aggregate total citations per year from citations collection and store in total_citations collection."""
    try:
        from db_config import get_collection
        citations_collection = get_collection('citations')
        total_citations_collection = get_collection('total_citations')
        
        # Ensure the collection exists by inserting a dummy document and then removing it
        try:
            total_citations_collection.insert_one({'dummy': 'dummy'})
            total_citations_collection.delete_one({'dummy': 'dummy'})
        except Exception as e:
            print(f"Error ensuring collection exists: {e}")
        
        all_citations = list(citations_collection.find({}))
        print(f"Found {len(all_citations)} citation documents to process")
        
        yearly_totals = {}
        for citation_doc in all_citations:
            citations_per_year = citation_doc.get('citationsPerYear', {})
            if hasattr(citations_per_year, 'to_dict'):
                citations_per_year = citations_per_year.to_dict()
            for year, count in citations_per_year.items():
                if str(year).isdigit():
                    y = int(year)
                    yearly_totals[y] = yearly_totals.get(y, 0) + int(count)
        
        print(f"Calculated totals for years: {list(yearly_totals.keys())}")
        
        # Clear existing data and insert new totals
        delete_result = total_citations_collection.delete_many({})
        print(f"Deleted {delete_result.deleted_count} existing documents")
        
        # Insert new totals
        insert_docs = [{'year': year, 'total': total} for year, total in yearly_totals.items()]
        if insert_docs:
            insert_result = total_citations_collection.insert_many(insert_docs)
            print(f"Inserted {len(insert_result.inserted_ids)} new documents")
        
        # Verify the data was stored
        stored_count = total_citations_collection.count_documents({})
        print(f"Total documents in total_citations collection: {stored_count}")
        
        return jsonify({
            'message': 'Total citations per year updated successfully.',
            'years': list(yearly_totals.keys()),
            'stored_count': stored_count
        }), 200
    except Exception as e:
        print(f"Error in update_total_citations: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/community/total_citations', methods=['GET'])
def get_total_citations():
    """Return total citations per year from total_citations collection as a year->total mapping."""
    try:
        from db_config import get_collection
        total_citations_collection = get_collection('total_citations')
        docs = list(total_citations_collection.find({}, {'_id': 0}))
        result = {str(doc['year']): doc['total'] for doc in docs}
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/community/all_citations', methods=['GET'])
def get_all_citations():
    """Return all teachers' citationsPerYear from the citations collection."""
    try:
        from db_config import get_collection
        citations_collection = get_collection('citations')
        docs = list(citations_collection.find({}, {'_id': 0, 'teacherName': 1, 'citationsPerYear': 1}))
        return jsonify({'teachers': docs}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/publication/details', methods=['GET'])
def get_publication_details_across_collections():
    """Search for a publication by title, year, and authors across all teacher collections. Returns the first match found."""
    try:
        from db_config import get_collection
        db = get_collection('teachers').database
        all_collections = db.list_collection_names()
        teacher_paper_collections = [col for col in all_collections if col.startswith('papers_')]

        # Get query parameters
        title = (request.args.get('title') or '').strip().lower()
        year = (request.args.get('year') or '').strip()
        authors = (request.args.get('authors') or '').strip().lower()
        if not title or not year or not authors:
            return jsonify({'error': 'title, year, and authors are required'}), 400

        for col_name in teacher_paper_collections:
            col = db.get_collection(col_name)
            for pub in col.find({}):
                pub_title = (pub.get('title', '') or '').strip().lower()
                pub_year = str(pub.get('year', '') or '').strip()
                pub_authors = (pub.get('authors', '') or '').strip().lower()
                if pub_title == title and pub_year == year and pub_authors == authors:
                    pub['_collection'] = col_name
                    return jsonify({'publication': pub}), 200
        return jsonify({'error': 'Publication not found'}), 404
    except Exception as error:
        return jsonify({'error': str(error)}), 500





@app.route('/admin/login', methods=['POST'])
def admin_login():
    from db_config import get_collection
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    admin_collection = get_collection('admin')
    # Remove legacy fallback: do NOT create a default admin if none exists
    admin = admin_collection.find_one({'username': username})
    if not admin:
        return jsonify({'error': 'Invalid username or password'}), 401
    if bcrypt.checkpw(password.encode('utf-8'), admin['password']):
        return jsonify({'success': True}), 200
    else:
        return jsonify({'error': 'Invalid username or password'}), 401

@app.route('/admin/change_password', methods=['POST'])
def admin_change_password():
    from db_config import get_collection
    data = request.get_json()
    username = data.get('username')
    old_password = data.get('oldPassword')
    new_password = data.get('newPassword')
    if not username or not old_password or not new_password:
        return jsonify({'error': 'All fields are required'}), 400
    admin_collection = get_collection('admin')
    admin = admin_collection.find_one({'username': username})
    if not admin:
        return jsonify({'error': 'User not found'}), 404
    if not bcrypt.checkpw(old_password.encode('utf-8'), admin['password']):
        return jsonify({'error': 'Old password is incorrect'}), 401
    hashed = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
    admin_collection.update_one({'username': username}, {'$set': {'password': hashed}})
    return jsonify({'success': True}), 200

@app.route('/admin/signup', methods=['POST'])
def admin_signup():
    from db_config import get_collection
    import secrets
    import string
    
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    answer1 = data.get('answer1')
    answer2 = data.get('answer2')
    security_code = data.get('securityCode', '')  # Optional for first admin
    
    if not username or not password or not answer1 or not answer2:
        return jsonify({'error': 'All fields are required'}), 400
    
    admin_collection = get_collection('admin')
    
    # Check if any admin already exists
    existing_admin = admin_collection.find_one({})
    
    if existing_admin:
        # Admin exists, require security code
        if not security_code:
            return jsonify({'error': 'Security code is required for admin replacement'}), 400
        
        # Check if security code matches
        if not bcrypt.checkpw(security_code.encode('utf-8'), existing_admin.get('security_code', b'')):
            return jsonify({'error': 'Invalid security code'}), 401
        
        # Check if username already exists (shouldn't happen with single admin, but good to check)
        if admin_collection.find_one({'username': username}):
            return jsonify({'error': 'Username already exists'}), 409
        
        # Delete the previous admin and create the new one
        admin_collection.delete_many({})  # Remove all existing admins
        
        # Generate a new secure security code for the new admin
        alphabet = string.ascii_uppercase + string.digits
        generated_security_code = ''.join(secrets.choice(alphabet) for _ in range(8))
        hashed_security_code = bcrypt.hashpw(generated_security_code.encode('utf-8'), bcrypt.gensalt())
        
        hashed_pw = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        hashed_a1 = bcrypt.hashpw(answer1.lower().strip().encode('utf-8'), bcrypt.gensalt())
        hashed_a2 = bcrypt.hashpw(answer2.lower().strip().encode('utf-8'), bcrypt.gensalt())
        
        admin_collection.insert_one({
            'username': username,
            'password': hashed_pw,
            'security': {
                'q1': "what's your favourite colour",
                'a1': hashed_a1,
                'q2': "where were you born",
                'a2': hashed_a2
            },
            'security_code': hashed_security_code,
            'is_primary_admin': True
        })
        
        return jsonify({
            'success': True, 
            'message': 'Admin replaced successfully',
            'security_code': generated_security_code,
            'note': 'Previous admin has been replaced. Please save this new security code securely.'
        }), 201
    
    else:
        # No admin exists, create the first admin
        if admin_collection.find_one({'username': username}):
            return jsonify({'error': 'Username already exists'}), 409
        
        # Generate a secure security code
        alphabet = string.ascii_uppercase + string.digits
        generated_security_code = ''.join(secrets.choice(alphabet) for _ in range(8))
        hashed_security_code = bcrypt.hashpw(generated_security_code.encode('utf-8'), bcrypt.gensalt())
        
        hashed_pw = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        hashed_a1 = bcrypt.hashpw(answer1.lower().strip().encode('utf-8'), bcrypt.gensalt())
        hashed_a2 = bcrypt.hashpw(answer2.lower().strip().encode('utf-8'), bcrypt.gensalt())
        
        admin_collection.insert_one({
            'username': username,
            'password': hashed_pw,
            'security': {
                'q1': "what's your favourite colour",
                'a1': hashed_a1,
                'q2': "where were you born",
                'a2': hashed_a2
            },
            'security_code': hashed_security_code,
            'is_primary_admin': True
        })
        
        return jsonify({
            'success': True, 
            'message': 'First admin created successfully',
            'security_code': generated_security_code,
            'note': 'Please save this security code securely. It will be required for any future admin replacements.'
        }), 201

@app.route('/admin/forgot_password', methods=['POST'])
def admin_forgot_password():
    from db_config import get_collection
    data = request.get_json()
    username = data.get('username')
    answer1 = data.get('answer1')
    answer2 = data.get('answer2')
    if not username or not answer1 or not answer2:
        return jsonify({'error': 'All fields are required'}), 400
    admin_collection = get_collection('admin')
    admin = admin_collection.find_one({'username': username})
    if not admin:
        return jsonify({'error': 'User not found'}), 404
    # Check answers (case-insensitive, trimmed)
    if not (bcrypt.checkpw(answer1.lower().strip().encode('utf-8'), admin['security']['a1']) and bcrypt.checkpw(answer2.lower().strip().encode('utf-8'), admin['security']['a2'])):
        return jsonify({'error': 'Security answers do not match'}), 401
    return jsonify({'success': True}), 200

@app.route('/admin/reset_password', methods=['POST'])
def admin_reset_password():
    from db_config import get_collection
    data = request.get_json()
    username = data.get('username')
    new_password = data.get('newPassword')
    if not username or not new_password:
        return jsonify({'error': 'All fields are required'}), 400
    admin_collection = get_collection('admin')
    admin = admin_collection.find_one({'username': username})
    if not admin:
        return jsonify({'error': 'User not found'}), 404
    hashed = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
    admin_collection.update_one({'username': username}, {'$set': {'password': hashed}})
    return jsonify({'success': True}), 200
    
@app.route('/admin/status', methods=['GET'])
def admin_status():
    """Check if admin exists and return status"""
    from db_config import get_collection
    admin_collection = get_collection('admin')
    admin_exists = admin_collection.count_documents({}) > 0
    
    return jsonify({
        'admin_exists': admin_exists,
        'message': 'Admin exists' if admin_exists else 'No admin found'
    }), 200

@app.route('/teachers/<teacher_id>/add_publication', methods=['POST'])
def add_publication_to_teacher(teacher_id):
    try:
        from db_config import get_collection
        import re, urllib.parse, datetime
        data = request.get_json(force=True)
        required_fields = ['teacherName', 'title']
        pub_type = data.get('publicationType')
        if pub_type == 'conference':
            required_fields.append('conference')
        elif pub_type == 'journal':
            required_fields.append('journal')
        elif pub_type == 'book':
            required_fields.append('book')
        elif pub_type == 'patent':
            required_fields += ['inventors']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400
        decoded_teacher_id = urllib.parse.unquote(teacher_id)
        collection_name = 'papers_' + re.sub(r'[^a-z0-9]', '_', decoded_teacher_id.lower())
        papers_collection = get_collection(collection_name)
        now = datetime.datetime.utcnow()
        doc = {
            'title': data['title'],
            'teacherName': data['teacherName'],
            'authors': data.get('authors', ''),
            'source': data.get('source', ''),
            'journal': data.get('journal', ''),
            'conference': data.get('conference', ''),
            'book': data.get('book', ''),
            'year': data.get('year', ''),
            'volume': data.get('volume', ''),
            'issue': data.get('issue', ''),
            'pages': data.get('pages', ''),
            'publisher': data.get('publisher', ''),
            'description': data.get('description', ''),
            'summary': data.get('summary', ''),
            'pdfLink': data.get('pdfLink', ''),
            'citationCount': int(data.get('citationCount', 0)),
            'publicationDate': data.get('publicationDate', ''),
            'inventors': data.get('inventors', ''),
            'patentOffice': data.get('patentOffice', ''),
            'patentNumber': data.get('patentNumber', ''),
            'applicationNumber': data.get('applicationNumber', ''),
            'patentFilingNumber': data.get('patentFilingNumber', ''),
            'filedOn': data.get('filedOn', ''),
            'grantedOn': data.get('grantedOn', ''),
            'patent': pub_type == 'patent',
            'createdAt': now,
            'updatedAt': now,
            '__v': 0
        }
        # Only add url if it is a non-empty, non-blank string and not None/null
        url = data.get('url', None)
        print(f"DEBUG: incoming url value: {repr(url)}")
        if url is not None and isinstance(url, str) and url.strip():
            doc['url'] = url.strip()
        result = papers_collection.insert_one(doc)
        doc['_id'] = str(result.inserted_id)
        return jsonify({'publication': doc}), 201
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/api/domains', methods=['GET'])
def get_domains():
    """Get all unique domains"""
    try:
        db = db_manager.connect()
        domains_collection = db['domains']
        
        # Get all unique domain names
        unique_domains = domains_collection.distinct('domainName')
        
        # Normalize and consolidate domain names
        domain_consolidation = {}
        domain_mapping = {}
        
        for domain_name in unique_domains:
            normalized_name = normalize_domain_name(domain_name)
            
            if normalized_name not in domain_consolidation:
                domain_consolidation[normalized_name] = {
                    'displayName': normalized_name,
                    'originalNames': [],
                    'teacherCount': 0
                }
            
            domain_consolidation[normalized_name]['originalNames'].append(domain_name)
            domain_mapping[domain_name] = normalized_name
        
        # Count teachers for each consolidated domain
        for domain_name in unique_domains:
            normalized_name = domain_mapping[domain_name]
            teacher_count = domains_collection.count_documents({'domainName': domain_name})
            domain_consolidation[normalized_name]['teacherCount'] += teacher_count
        
        # Convert to list format
        domain_stats = []
        for normalized_name, data in domain_consolidation.items():
            domain_stats.append({
                'domainName': data['displayName'],
                'teacherCount': data['teacherCount'],
                'originalNames': data['originalNames']
            })
        
        # Sort by teacher count (descending)
        domain_stats.sort(key=lambda x: x['teacherCount'], reverse=True)
        
        return jsonify({
            'domains': domain_stats,
            'total_domains': len(domain_stats)
        }), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

def normalize_domain_name(domain_name):
    """Normalize domain names to consolidate similar variants"""
    if not domain_name:
        return domain_name
    
    # Convert to lowercase and remove extra spaces
    normalized = domain_name.lower().strip()
    
    # Cybersecurity variants
    if any(variant in normalized for variant in ['cyber security', 'cybersecurity', 'cyber-security']):
        return 'Cybersecurity'
    
    # Machine Learning variants
    if any(variant in normalized for variant in ['machine learning', 'ml', 'ai/ml', 'artificial intelligence']):
        return 'Machine Learning'
    
    # Digital Twin variants
    if any(variant in normalized for variant in ['digital twin', 'digital twins']):
        return 'Digital Twin'
    
    # Digital Forensics variants
    if any(variant in normalized for variant in ['digital forensics', 'cyber forensics', 'computer forensics']):
        return 'Digital Forensics'
    
    # Computer Vision variants
    if any(variant in normalized for variant in ['computer vision', 'cv', 'image processing']):
        return 'Computer Vision'
    
    # Data Science variants
    if any(variant in normalized for variant in ['data science', 'data analytics', 'big data']):
        return 'Data Science'
    
    # Internet of Things variants
    if any(variant in normalized for variant in ['iot', 'internet of things', 'smart devices']):
        return 'Internet of Things'
    
    # Blockchain variants
    if any(variant in normalized for variant in ['blockchain', 'distributed ledger', 'cryptocurrency']):
        return 'Blockchain'
    
    # Cloud Computing variants
    if any(variant in normalized for variant in ['cloud computing', 'cloud', 'aws', 'azure']):
        return 'Cloud Computing'
    
    # If no match found, return the original name with proper capitalization
    return domain_name.title()

@app.route('/api/domains/<path:domain_name>/teachers', methods=['GET'])
def get_teachers_by_domain(domain_name):
    """Get all teachers for a specific domain"""
    try:
        db = db_manager.connect()
        domains_collection = db['domains']
        
        # Decode the domain name from URL
        decoded_domain_name = urllib.parse.unquote(domain_name)
        
        # Get all original domain names that map to this consolidated domain
        all_domains = domains_collection.distinct('domainName')
        matching_original_names = []
        
        # Normalize requested name too to compare canonical forms
        normalized_requested = normalize_domain_name(decoded_domain_name)

        for original_name in all_domains:
            if normalize_domain_name(original_name) == normalized_requested:
                matching_original_names.append(original_name)
        
        # Find all teachers in any of the matching original domain names
        teachers = []
        for original_name in matching_original_names:
            domain_entries = domains_collection.find({'domainName': original_name})
            for entry in domain_entries:
                teachers.append({
                    'teacherName': entry['teacherName'],
                    'domainName': entry['domainName'],
                    'domainUrl': entry.get('domainUrl', ''),
                    'lastUpdated': entry.get('lastUpdated', '').isoformat() if entry.get('lastUpdated') else None
                })
        
        # Remove duplicate teachers (in case a teacher has multiple variants of the same domain)
        unique_teachers = []
        seen_teachers = set()
        for teacher in teachers:
            if teacher['teacherName'] not in seen_teachers:
                unique_teachers.append(teacher)
                seen_teachers.add(teacher['teacherName'])
        
        return jsonify({
            'domainName': decoded_domain_name,
            'teachers': unique_teachers,
            'teacherCount': len(unique_teachers)
        }), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

# Yearly Projects API endpoints
@app.route('/api/yearly-projects', methods=['GET'])
def get_yearly_projects():
    """Get all yearly projects"""
    try:
        db = db_manager.connect()
        projects_collection = db['yearly_projects']
        
        # Get all projects sorted by year (descending) and creation date (descending)
        # Explicitly include _id field in the query
        projects = list(projects_collection.find({}, {'_id': 1, 'year': 1, 'teacherName': 1, 'projectName': 1, 'projectDescription': 1, 'students': 1, 'createdAt': 1, 'report': 1, 'poster': 1}).sort([
            ('year', -1), 
            ('createdAt', -1)
        ]))
        
        print(f"Found {len(projects)} projects in database")
        
        # Convert ObjectId to string for JSON serialization
        for i, project in enumerate(projects):
            print(f"Project {i}: {project}")
            if '_id' in project:
                project['_id'] = str(project['_id'])
                print(f"Converted _id to string: {project['_id']}")
            else:
                print(f"Project {i} missing _id field!")
                # Try to get the _id from the document
                project_doc = projects_collection.find_one({
                    'year': project['year'],
                    'teacherName': project['teacherName'],
                    'projectName': project['projectName']
                })
                if project_doc and '_id' in project_doc:
                    project['_id'] = str(project_doc['_id'])
                    print(f"Retrieved _id from database: {project['_id']}")
                else:
                    print(f"Could not retrieve _id for project {i}")
        
        return jsonify({
            'success': True,
            'projects': projects
        }), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/api/yearly-projects/bulk', methods=['POST'])
def add_yearly_projects_bulk():
    """Upload an Excel file and create multiple yearly projects grouped by group_id.

    Expected columns: group_id, name, srn, mentor, project_title, project_description, year, report, poster
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Empty filename'}), 400

        # Read workbook in memory
        data = file.read()
        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active

        # Extract headers
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'error': 'Empty sheet'}), 400
        headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

        required = ['group_id', 'name', 'srn', 'mentor', 'project_title', 'project_description', 'year', 'report', 'poster']
        for req in required:
            if req not in headers:
                return jsonify({'error': f'Missing required column: {req}'}), 400

        idx = {h: headers.index(h) for h in required}

        # Group rows by group_id
        groups = {}
        for r in rows[1:]:
            if r is None:
                continue
            group_id = str(r[idx['group_id']]).strip() if r[idx['group_id']] is not None else ''
            if not group_id:
                # Skip rows without a group id
                continue
            group = groups.setdefault(group_id, {
                'mentor': '',
                'project_title': '',
                'project_description': '',
                'year': '',
                'report': '',
                'poster': '',
                'students': []
            })

            mentor = (str(r[idx['mentor']]) if r[idx['mentor']] is not None else '').strip()
            project_title = (str(r[idx['project_title']]) if r[idx['project_title']] is not None else '').strip()
            project_description = (str(r[idx['project_description']]) if r[idx['project_description']] is not None else '').strip()
            year_val = r[idx['year']]
            try:
                year = int(str(year_val).strip()) if year_val is not None else ''
            except Exception:
                year = ''
            report = (str(r[idx['report']]) if r[idx['report']] is not None else '').strip()
            poster = (str(r[idx['poster']]) if r[idx['poster']] is not None else '').strip()
            name = (str(r[idx['name']]) if r[idx['name']] is not None else '').strip()
            srn = (str(r[idx['srn']]) if r[idx['srn']] is not None else '').strip()

            # Set shared fields if present (prefer first non-empty)
            if project_title and not group['project_title']:
                group['project_title'] = project_title
            if mentor and not group['mentor']:
                group['mentor'] = mentor
            if project_description and not group['project_description']:
                group['project_description'] = project_description
            if year and not group['year']:
                group['year'] = year
            if report and not group['report']:
                group['report'] = report
            if poster and not group['poster']:
                group['poster'] = poster

            if name or srn:
                group['students'].append({'name': name, 'srn': srn})

        # Persist projects
        from db_config import get_collection
        projects_collection = get_collection('yearly_projects')

        # Create unique index (no-op if exists)
        try:
            projects_collection.create_index([
                ('year', 1), ('teacherName', 1), ('projectName', 1)
            ], unique=True)
        except Exception:
            pass

        inserted = []
        skipped = []
        preview = []
        for gid, g in groups.items():
            doc = {
                'year': g['year'] if isinstance(g['year'], int) else int(str(g['year']) or 0),
                'teacherName': g['mentor'],
                'projectName': g['project_title'],
                'projectDescription': g['project_description'],
                'students': g['students'],
                'report': g['report'],
                'poster': g['poster'],
                'createdAt': datetime.utcnow().isoformat(),
                'groupId': gid
            }
            # Skip duplicates
            existing = projects_collection.find_one({
                'year': doc['year'],
                'teacherName': doc['teacherName'],
                'projectName': doc['projectName']
            })
            if existing:
                skipped.append({'groupId': gid, 'reason': 'duplicate', 'projectId': str(existing.get('_id'))})
                continue
            res = projects_collection.insert_one(doc)
            inserted.append(str(res.inserted_id))
            preview.append({
                '_id': str(res.inserted_id),
                'year': doc['year'],
                'teacherName': doc['teacherName'],
                'projectName': doc['projectName'],
                'projectDescription': doc['projectDescription'],
                'students': doc['students'],
                'report': doc['report'],
                'poster': doc['poster']
            })

        return jsonify({
            'success': True,
            'inserted': inserted,
            'skipped': skipped,
            'preview': preview,
            'groups': len(groups)
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/api/yearly-projects', methods=['POST'])
def add_yearly_project():
    """Add a new yearly project"""
    try:
        db = db_manager.connect()
        projects_collection = db['yearly_projects']
        
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['year', 'teacherName', 'projectName', 'projectDescription']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Create project document
        project = {
            'year': int(data['year']),
            'teacherName': data['teacherName'],
            'projectName': data['projectName'],
            'projectDescription': data['projectDescription'],
            'students': data.get('students'),
            'createdAt': datetime.utcnow().isoformat()
        }
        
        # Ensure a unique index to help prevent duplicates (year, teacherName, projectName)
        try:
            projects_collection.create_index([
                ('year', 1), ('teacherName', 1), ('projectName', 1)
            ], unique=True)
        except Exception:
            pass

        # Prevent duplicates
        existing = projects_collection.find_one({
            'year': project['year'],
            'teacherName': project['teacherName'],
            'projectName': project['projectName']
        })
        if existing:
            return jsonify({'error': 'Duplicate project exists', 'projectId': str(existing.get('_id'))}), 409
        
        # Insert the project
        result = projects_collection.insert_one(project)
        return jsonify({
            'success': True,
            'message': 'Project added successfully',
            'projectId': str(result.inserted_id)
        }), 201
            
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/api/yearly-projects/<project_id>', methods=['PUT'])
def update_yearly_project(project_id):
    """Update an existing yearly project"""
    try:
        db = db_manager.connect()
        projects_collection = db['yearly_projects']
        
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['year', 'teacherName', 'projectName', 'projectDescription']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Update project document
        update_data = {
            'year': int(data['year']),
            'teacherName': data['teacherName'],
            'projectName': data['projectName'],
            'projectDescription': data['projectDescription'],
            'students': data.get('students'),
            'updatedAt': datetime.utcnow().isoformat()
        }
        
        # Update the project
        result = projects_collection.update_one(
            {'_id': ObjectId(project_id)},
            {'$set': update_data}
        )
        
        if result.modified_count > 0:
            return jsonify({
                'success': True,
                'message': 'Project updated successfully'
            }), 200
        else:
            return jsonify({'error': 'Project not found or no changes made'}), 404
            
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/api/test-patents', methods=['GET'])
def test_patents_endpoint():
    """Test endpoint to check patent data and counting"""
    try:
        from db_config import get_collection
        db = get_collection('teachers').database
        all_collections = db.list_collection_names()
        teacher_paper_collections = [col for col in all_collections if col.startswith('papers_')]
        
        patent_data = []
        total_patents = 0
        
        for col_name in teacher_paper_collections:
            col = db.get_collection(col_name)
            # Find all documents with patent field
            patents = list(col.find({'patent': {'$exists': True}}))
            
            for patent in patents:
                patent_info = {
                    'collection': col_name,
                    'title': patent.get('title', 'No title'),
                    'patent_field': patent.get('patent'),
                    'patent_type': type(patent.get('patent')).__name__,
                    'year': patent.get('year'),
                    'teacherName': patent.get('teacherName')
                }
                patent_data.append(patent_info)
                
                # Check if it would be counted by our logic
                patent_value = patent.get('patent')
                is_counted = (
                    patent_value is True or 
                    patent_value == True or 
                    (isinstance(patent_value, str) and patent_value.lower() == 'true')
                )
                
                if is_counted:
                    total_patents += 1
                    patent_info['counted'] = True
                else:
                    patent_info['counted'] = False
        
        return jsonify({
            'success': True,
            'total_patents_found': len(patent_data),
            'total_patents_counted': total_patents,
            'patent_details': patent_data
        }), 200
        
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/api/test-delete', methods=['GET'])
def test_delete_endpoint():
    """Test endpoint to verify backend is working"""
    try:
        db = db_manager.connect()
        projects_collection = db['yearly_projects']
        count = projects_collection.count_documents({})
        return jsonify({
            'success': True,
            'message': f'Backend is working. Found {count} projects in collection.',
            'collection_name': 'yearly_projects'
        }), 200
    except Exception as error:
        return jsonify({'error': str(error)}), 500

@app.route('/api/yearly-projects/<project_id>', methods=['DELETE'])
def delete_yearly_project(project_id):
    """Delete a yearly project"""
    try:
        print(f"Attempting to delete project with ID: {project_id}")
        
        # Validate project_id format
        if not project_id or len(project_id) != 24:
            return jsonify({'error': 'Invalid project ID format'}), 400
        
        db = db_manager.connect()
        projects_collection = db['yearly_projects']
        
        # First check if project exists
        project = projects_collection.find_one({'_id': ObjectId(project_id)})
        if not project:
            print(f"Project not found with ID: {project_id}")
            return jsonify({'error': 'Project not found'}), 404
        
        print(f"Found project: {project.get('projectName', 'Unknown')} - {project.get('teacherName', 'Unknown')}")
        
        # Delete the project
        result = projects_collection.delete_one({'_id': ObjectId(project_id)})
        
        if result.deleted_count > 0:
            print(f"Successfully deleted project with ID: {project_id}")
            return jsonify({
                'success': True,
                'message': 'Project deleted successfully'
            }), 200
        else:
            print(f"Failed to delete project with ID: {project_id}")
            return jsonify({'error': 'Failed to delete project'}), 500
            
    except Exception as error:
        print(f"Error deleting project: {str(error)}")
        return jsonify({'error': str(error)}), 500

# Start the Node.js scheduler as a background process when the Flask app starts
# Only do this if this script is the main entry point and Redis is configured
if __name__ == '__main__':
    if os.getenv('REDIS_URL') or os.getenv('UPSTASH_REDIS_URL') or os.getenv('REDIS_HOST'):
        scheduler_path = os.path.join(os.path.dirname(__file__), 'scheduler.js')
        try:
            subprocess.Popen(['node', scheduler_path], cwd=os.path.dirname(__file__))
            print('Started scheduler.js as a background process.')
        except Exception as e:
            print(f'Failed to start scheduler.js: {e}')
    else:
        print('Scheduler not started: set REDIS_URL or REDIS_HOST to enable background jobs.')

if __name__ == '__main__':
    # Prefer Railway's PORT, fallback to API_PORT, then 5000 locally
    port = int(os.getenv('PORT', os.getenv('API_PORT', 5000)))
    # Disable debug mode and auto-reloader to prevent conflicts with Celery
    app.run(host='0.0.0.0', port=port, debug=False, use_reloader=False)